from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Cronograma"

WEEKS = 16

# ── Colores ────────────────────────────────────────────────────────────────────
BLUE_FILL   = PatternFill("solid", fgColor="4472C4")   # azul OE1
GREEN_FILL  = PatternFill("solid", fgColor="70AD47")   # verde OE2
HDR_FILL    = PatternFill("solid", fgColor="1F3864")   # azul marino encabezado
OE1_HDR     = PatternFill("solid", fgColor="2E75B6")   # azul medio fila OE1
OE2_HDR     = PatternFill("solid", fgColor="375623")   # verde oscuro fila OE2
GRAY_FILL   = PatternFill("solid", fgColor="F2F2F2")   # gris claro celda vacía
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")

# ── Bordes ─────────────────────────────────────────────────────────────────────
thin  = Side(style="thin",   color="BFBFBF")
thick = Side(style="medium", color="595959")

def cell_border(top=thin, bottom=thin, left=thin, right=thin):
    return Border(top=top, bottom=bottom, left=left, right=right)

border_normal = cell_border()
border_thick  = cell_border(top=thick, bottom=thick, left=thick, right=thick)

# ── Anchos de columna ─────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 44
for col in range(2, 2 + WEEKS):
    ws.column_dimensions[get_column_letter(col)].width = 4.2

# ── Fila de encabezado (semanas) ──────────────────────────────────────────────
ws.row_dimensions[1].height = 20
ws["A1"].value = "Actividades"
ws["A1"].fill  = HDR_FILL
ws["A1"].font  = Font(bold=True, color="FFFFFF", size=10)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws["A1"].border = border_normal

for w in range(1, WEEKS + 1):
    col = w + 1
    cell = ws.cell(row=1, column=col, value=f"S{w}")
    cell.fill = HDR_FILL
    cell.font = Font(bold=True, color="FFFFFF", size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_normal

# ── Datos ─────────────────────────────────────────────────────────────────────
# (label, start_week, end_week, tipo)
# tipo: "hdr1", "hdr2", "oe1", "oe2"
rows = [
    ("Objetivo Específico 1",                                          None, None, "hdr1"),
    ("1.1  Recopilación y estructuración del catálogo alimentario",      1,    4,  "oe1"),
    ("1.2  Definición arquitectura C4 y configuración pipeline RAG",     3,    7,  "oe1"),
    ("1.3  Desarrollo de la interfaz del sistema",                       6,   11,  "oe1"),
    ("Objetivo Específico 2",                                          None, None, "hdr2"),
    ("2.1  Diseño y validación del cuestionario Likert",                 9,   12,  "oe2"),
    ("2.2  Definición y ejecución de casos de prueba",                  11,   14,  "oe2"),
    ("2.3  Aplicación del cuestionario y análisis estadístico",         13,   16,  "oe2"),
]

for i, (label, start, end, rtype) in enumerate(rows):
    excel_row = i + 2
    ws.row_dimensions[excel_row].height = 22

    # ── Celda de etiqueta ─────────────────────────────────────────────────────
    cell = ws.cell(row=excel_row, column=1, value=label)
    cell.alignment = Alignment(horizontal="left", vertical="center",
                               indent=0 if rtype in ("hdr1","hdr2") else 1,
                               wrap_text=False)
    cell.border = border_normal

    if rtype == "hdr1":
        cell.fill = OE1_HDR
        cell.font = Font(bold=True, color="FFFFFF", size=10)
    elif rtype == "hdr2":
        cell.fill = OE2_HDR
        cell.font = Font(bold=True, color="FFFFFF", size=10)
    else:
        cell.fill = WHITE_FILL
        cell.font = Font(size=9, color="404040")

    # ── Celdas de semanas ─────────────────────────────────────────────────────
    for w in range(1, WEEKS + 1):
        col = w + 1
        wc = ws.cell(row=excel_row, column=col, value="")
        wc.border = border_normal
        wc.alignment = Alignment(horizontal="center", vertical="center")

        if rtype in ("hdr1", "hdr2"):
            wc.fill = OE1_HDR if rtype == "hdr1" else OE2_HDR
        elif start and end and start <= w <= end:
            wc.fill = BLUE_FILL if rtype == "oe1" else GREEN_FILL
        else:
            wc.fill = GRAY_FILL

# ── Borde exterior grueso ─────────────────────────────────────────────────────
total_rows = len(rows) + 1
total_cols = WEEKS + 1

for r in range(1, total_rows + 1):
    for c in range(1, total_cols + 1):
        cell = ws.cell(row=r, column=c)
        top    = thick if r == 1             else thin
        bottom = thick if r == total_rows    else thin
        left   = thick if c == 1             else thin
        right  = thick if c == total_cols    else thin
        cell.border = Border(top=top, bottom=bottom, left=left, right=right)

# ── Guardar ───────────────────────────────────────────────────────────────────
out = r"c:\Users\ASUS\Downloads\ClaudeProyecto\cronograma.xlsx"
wb.save(out)
print("Guardado:", out)
